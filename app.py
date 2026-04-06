"""
MEAL Co-Pilot — Momentum Incubator
Beautiful Dash application with slide-based views and dark navy theme.

Run with:
    pip install dash dash-bootstrap-components openai
    python meal_copilot_app.py

The app calls the OpenAI API (uses OPENAI_API_KEY env variable).
"""

import dash
from dash import dcc, html, Input, Output, State, callback_context, no_update
import dash_bootstrap_components as dbc
import json
import os
import io
import base64
import pathlib
from typing import Dict, Any
from openai import OpenAI
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────
# CSS — written to assets/ so Dash auto-serves it
# ─────────────────────────────────────────────────────────────────
CUSTOM_CSS = """
:root {
  --navy-900: #050d1f;
  --navy-800: #0a1628;
  --navy-700: #0f2040;
  --navy-600: #152855;
  --navy-500: #1c3468;
  --accent:   #4f9cf7;
  --accent2:  #63e5ff;
  --gold:     #f0c040;
  --white:    #f0f4ff;
  --muted:    #8ba4c8;
  --card-bg:  rgba(15, 32, 64, 0.8);
  --border:   rgba(79, 156, 247, 0.2);
}
* { box-sizing: border-box; margin: 0; padding: 0; }
body {
  background: var(--navy-900);
  color: var(--white);
  font-family: 'DM Sans', sans-serif;
  min-height: 100vh;
  overflow-x: hidden;
}
.bg-mesh {
  position: fixed; inset: 0; z-index: 0; pointer-events: none;
  background:
    radial-gradient(ellipse 80% 60% at 20% 20%, rgba(79,156,247,0.08) 0%, transparent 60%),
    radial-gradient(ellipse 60% 80% at 80% 80%, rgba(99,229,255,0.06) 0%, transparent 60%),
    linear-gradient(180deg, var(--navy-900) 0%, var(--navy-800) 100%);
}
.grid-lines {
  position: fixed; inset: 0; z-index: 0; pointer-events: none;
  background-image:
    linear-gradient(rgba(79,156,247,0.04) 1px, transparent 1px),
    linear-gradient(90deg, rgba(79,156,247,0.04) 1px, transparent 1px);
  background-size: 60px 60px;
}
.slide-wrap {
  position: relative; z-index: 1;
  min-height: 100vh;
  display: flex; flex-direction: column;
}
.app-header {
  display: flex; align-items: center; justify-content: space-between;
  padding: 20px 48px;
  border-bottom: 1px solid var(--border);
  background: rgba(5, 13, 31, 0.7);
  backdrop-filter: blur(12px);
  position: sticky; top: 0; z-index: 100;
}
.logo-mark {
  font-family: 'Playfair Display', serif;
  font-size: 1.5rem; font-weight: 900;
  background: linear-gradient(135deg, var(--accent) 0%, var(--accent2) 100%);
  -webkit-background-clip: text; -webkit-text-fill-color: transparent;
  letter-spacing: -0.02em;
}
.logo-sub {
  font-size: 0.7rem; color: var(--muted);
  letter-spacing: 0.18em; text-transform: uppercase;
  margin-top: 2px;
}
.step-indicator { display: flex; gap: 8px; align-items: center; }
.step-dot {
  width: 8px; height: 8px; border-radius: 50%;
  background: var(--navy-500); border: 1px solid var(--border);
  transition: all 0.3s ease;
}
.step-dot.active {
  background: var(--accent);
  box-shadow: 0 0 12px rgba(79,156,247,0.6);
  border-color: var(--accent);
}
.step-dot.done { background: rgba(79,156,247,0.4); border-color: rgba(79,156,247,0.4); }
.slide {
  flex: 1; display: flex; flex-direction: column;
  justify-content: center; align-items: center;
  padding: 60px 48px;
  animation: slideIn 0.5s cubic-bezier(0.22, 1, 0.36, 1);
}
@keyframes slideIn {
  from { opacity: 0; transform: translateY(24px); }
  to   { opacity: 1; transform: translateY(0); }
}
.slide-title {
  font-family: 'Playfair Display', serif;
  font-size: clamp(2.2rem, 5vw, 3.6rem);
  font-weight: 900; line-height: 1.1;
  text-align: center; margin-bottom: 12px;
}
.slide-subtitle {
  color: var(--muted); font-size: 1.05rem;
  text-align: center; max-width: 540px;
  line-height: 1.6; margin-bottom: 48px;
}
.gradient-text {
  background: linear-gradient(135deg, var(--accent) 0%, var(--accent2) 100%);
  -webkit-background-clip: text; -webkit-text-fill-color: transparent;
}
.stat-strip {
  display: flex; gap: 40px; align-items: center;
  padding: 24px 40px; margin-bottom: 48px;
  background: var(--card-bg); border-radius: 16px;
  border: 1px solid var(--border);
}
.stat-divider { width: 1px; height: 40px; background: var(--border); }
.stat-num { font-size: 2rem; font-weight: 700; }
.stat-lbl { font-size: 0.78rem; color: var(--muted); text-transform: uppercase; letter-spacing: 0.1em; }
.form-card {
  background: var(--card-bg);
  border: 1px solid var(--border);
  border-radius: 20px; padding: 48px;
  width: 100%; max-width: 860px;
  backdrop-filter: blur(16px);
  box-shadow: 0 32px 80px rgba(0,0,0,0.4);
}
.form-row { display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 20px; }
.form-row.single { grid-template-columns: 1fr; }
.field-label { font-size: 0.72rem; letter-spacing: 0.12em; text-transform: uppercase; color: var(--accent); font-weight: 600; margin-bottom: 8px; }
.field-hint { font-size: 0.75rem; color: var(--muted); margin-bottom: 8px; }
input[type=text], textarea {
  background: rgba(5, 13, 31, 0.7) !important;
  border: 1px solid var(--border) !important;
  border-radius: 10px !important;
  color: var(--white) !important;
  font-family: 'DM Sans', sans-serif !important;
  font-size: 0.9rem !important;
  padding: 12px 16px !important;
  width: 100% !important;
  transition: border-color 0.2s, box-shadow 0.2s !important;
  outline: none !important;
}
input[type=text]:focus, textarea:focus {
  border-color: var(--accent) !important;
  box-shadow: 0 0 0 3px rgba(79,156,247,0.15) !important;
}
textarea { min-height: 88px !important; resize: vertical !important; }
.btn-primary {
  background: linear-gradient(135deg, var(--accent) 0%, #2563eb 100%);
  border: none; border-radius: 12px;
  color: #fff; cursor: pointer;
  font-family: 'DM Sans', sans-serif;
  font-size: 0.95rem; font-weight: 600;
  padding: 16px 40px; letter-spacing: 0.02em;
  transition: all 0.25s ease;
  box-shadow: 0 8px 32px rgba(79,156,247,0.3);
}
.btn-primary:hover { transform: translateY(-2px); box-shadow: 0 14px 40px rgba(79,156,247,0.45); }
.btn-secondary {
  background: transparent;
  border: 1px solid var(--border); border-radius: 12px;
  color: var(--muted); cursor: pointer;
  font-family: 'DM Sans', sans-serif;
  font-size: 0.9rem; font-weight: 500;
  padding: 14px 32px; transition: all 0.25s ease;
}
.btn-secondary:hover { border-color: var(--accent); color: var(--accent); }
.progress-wrap { text-align: center; max-width: 560px; }
.progress-icon { font-size: 3rem; margin-bottom: 24px; animation: pulse 1.8s ease-in-out infinite; }
@keyframes pulse { 0%,100%{transform:scale(1);opacity:1;} 50%{transform:scale(1.1);opacity:0.7;} }
.progress-bar-bg { background: var(--navy-700); border-radius: 8px; height: 6px; overflow: hidden; margin: 24px 0; }
.progress-bar-fill {
  height: 100%; border-radius: 8px;
  background: linear-gradient(90deg, var(--accent), var(--accent2), var(--accent));
  background-size: 200% 100%;
  animation: shimmer 2s linear infinite;
  transition: width 0.8s cubic-bezier(0.22,1,0.36,1);
}
@keyframes shimmer { 0%{background-position:200% 0;} 100%{background-position:-200% 0;} }
.progress-step { color: var(--muted); font-size: 0.9rem; display: flex; align-items: center; justify-content: center; gap: 8px; }
.results-wrap { width: 100%; max-width: 1100px; }
.tab-nav { display: flex; gap: 4px; flex-wrap: wrap; margin-bottom: 28px; justify-content: center; }
.tab-btn {
  background: transparent; border: 1px solid var(--border); border-radius: 8px;
  color: var(--muted); cursor: pointer; font-family: 'DM Sans', sans-serif;
  font-size: 0.78rem; font-weight: 500; padding: 8px 16px;
  letter-spacing: 0.04em; text-transform: uppercase; transition: all 0.2s;
}
.tab-btn:hover { border-color: var(--accent); color: var(--accent); }
.tab-btn.active { background: var(--accent); border-color: var(--accent); color: #fff; box-shadow: 0 4px 16px rgba(79,156,247,0.35); }
.result-card {
  background: var(--card-bg); border: 1px solid var(--border);
  border-radius: 16px; padding: 36px; backdrop-filter: blur(16px);
  animation: slideIn 0.35s ease;
}
.result-section-title {
  font-family: 'Playfair Display', serif; font-size: 1.4rem; font-weight: 700;
  color: var(--accent2); margin-bottom: 20px; padding-bottom: 12px;
  border-bottom: 1px solid var(--border);
}
.result-label {
  font-size: 0.7rem; letter-spacing: 0.14em; text-transform: uppercase;
  color: var(--accent); font-weight: 600; margin-bottom: 6px; margin-top: 18px;
}
.result-value { color: var(--white); font-size: 0.92rem; line-height: 1.6; }
.result-value ul { padding-left: 18px; }
.result-value li { margin-bottom: 4px; }
.badge-grid { display: flex; flex-wrap: wrap; gap: 8px; margin-top: 4px; }
.badge { background: rgba(79,156,247,0.12); border: 1px solid rgba(79,156,247,0.25); border-radius: 6px; color: var(--accent); font-size: 0.78rem; padding: 4px 10px; }
.badge.high { background: rgba(240,192,64,0.12); border-color: rgba(240,192,64,0.25); color: var(--gold); }
.badge.low  { background: rgba(100,200,120,0.12); border-color: rgba(100,200,120,0.25); color: #7ce0a0; }
.sh-table { width: 100%; border-collapse: collapse; margin-top: 12px; }
.sh-table th { font-size: 0.68rem; letter-spacing: 0.1em; text-transform: uppercase; color: var(--muted); padding: 10px 12px; text-align: left; border-bottom: 1px solid var(--border); }
.sh-table td { padding: 10px 12px; font-size: 0.85rem; color: var(--white); border-bottom: 1px solid rgba(79,156,247,0.06); }
.sh-table tr:hover td { background: rgba(79,156,247,0.04); }
.lf-table { width: 100%; border-collapse: collapse; }
.lf-table th { font-size: 0.68rem; letter-spacing: 0.12em; text-transform: uppercase; color: var(--accent); padding: 14px; text-align: left; background: rgba(79,156,247,0.08); border: 1px solid var(--border); }
.lf-table td { padding: 14px; font-size: 0.84rem; color: var(--white); border: 1px solid var(--border); vertical-align: top; line-height: 1.5; }
.lf-table .lc { font-weight: 700; color: var(--accent2); background: rgba(15,32,64,0.6); white-space: nowrap; }
.strategy-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(240px,1fr)); gap: 16px; margin-top: 12px; }
.strategy-card { background: rgba(5,13,31,0.5); border: 1px solid var(--border); border-radius: 12px; padding: 20px; transition: all 0.2s; }
.strategy-card:hover { border-color: var(--accent); transform: translateY(-2px); }
.strategy-card.rec { border-color: var(--gold); background: rgba(240,192,64,0.04); }
.score-bar { height: 4px; border-radius: 4px; background: var(--navy-700); margin-top: 8px; }
.score-fill { height: 100%; border-radius: 4px; background: linear-gradient(90deg, var(--accent), var(--accent2)); }
.error-box { background: rgba(220,60,60,0.08); border: 1px solid rgba(220,60,60,0.3); border-radius: 12px; padding: 20px; color: #ff9090; font-size: 0.9rem; }
.highlight-box { padding: 16px; background: rgba(79,156,247,0.06); border-radius: 10px; border: 1px solid var(--border); margin-top: 6px; }
.gold-box { padding: 16px; background: rgba(240,192,64,0.06); border-radius: 10px; border: 1px solid rgba(240,192,64,0.2); margin-top: 6px; }
.cyan-box { padding: 14px; background: rgba(99,229,255,0.06); border-radius: 10px; border: 1px solid rgba(99,229,255,0.15); margin-top: 6px; }
"""

_here = pathlib.Path(__file__).parent
_assets = _here / "assets"
_assets.mkdir(exist_ok=True)
(_assets / "style.css").write_text(CUSTOM_CSS)

# ─────────────────────────────────────────────────────────────────
# APP
# ─────────────────────────────────────────────────────────────────
app = dash.Dash(
    __name__,
    external_stylesheets=[
        dbc.themes.BOOTSTRAP,
        "https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;700;900&family=DM+Sans:wght@300;400;500;600&display=swap",
    ],
    suppress_callback_exceptions=True,
    title="MEAL Co-Pilot | Momentum Incubator",
    assets_folder=str(_assets),
)

# ─────────────────────────────────────────────────────────────────
# MEAL ENGINE
# ─────────────────────────────────────────────────────────────────
def run_meal_analysis(form_data: Dict) -> Dict:
    client = OpenAI()

    def call(prompt: str) -> dict:
        try:
            r = client.responses.create(
                model="gpt-5.4",
                input=[
                    {"role": "system", "content": "You are a MEAL expert. Always respond with valid JSON only, no markdown, no other text."},
                    {"role": "user", "content": prompt},
                ],
                max_output_tokens=2000,
            )
            text = r.output_text.strip()
            if text.startswith("```"):
                text = "\n".join(text.split("\n")[1:])
                text = text.rstrip("`").strip()
            return json.loads(text)
        except Exception as e:
            return {"error": str(e)}

    fd = json.dumps(form_data, indent=2)

    needs = call(f"""Perform a NEEDS ANALYSIS.
Project Data: {fd}
Return JSON:
{{"needs_statement":"string","maslow_analysis":{{"physical_needs":[],"safety_needs":[],"social_belonging":[],"esteem_needs":[]}},"needs_vs_wants":{{"critical_needs":[],"expressed_wants":[],"demands":[]}},"prioritized_needs":[{{"rank":1,"need":"string","justification":"string"}}]}}""")

    stakeholder = call(f"""Perform a STAKEHOLDER ANALYSIS.
Project Data: {fd}
Return JSON:
{{"stakeholder_matrix":[{{"name":"string","type":"Primary/Secondary/Tertiary","power":"High/Low","interest":"High/Low","grid_position":"Manage Closely/Keep Satisfied/Keep Informed/Monitor","engagement_strategy":"string"}}],"salience_classification":{{"definitive":[],"dependent":[],"latent":[]}},"vulnerable_groups":[]}}""")

    gdsi = call(f"""Perform a GENDER, DISABILITY & SOCIAL INCLUSION (GDSI) analysis.
Project Data: {fd}
Return JSON:
{{"excluded_groups":[],"barriers_analysis":[{{"group":"string","barrier_type":"Physical/Social/Communication/Institutional","description":"string"}}],"intersectional_risks":[],"do_no_harm_assessment":"string","inclusion_actions":{{"targeted_actions":[],"mainstreamed_actions":[]}}}}""")

    org = call(f"""Perform an ORGANIZATIONAL NEEDS ANALYSIS.
Project Data: {fd}
Return JSON:
{{"strategic_alignment":{{"org_mission":"string","project_contribution":"string","alignment_score":"High/Medium/Low"}},"performance_gaps":[{{"gap":"string","root_cause":"string","impact":"string"}}],"organizational_constraints":{{"budget":"string","personnel":"string","infrastructure":"string","culture_policies":"string"}},"change_readiness":"High/Medium/Low"}}""")

    problem = call(f"""Perform a PROBLEM TREE ANALYSIS.
Project Data: {fd}
Needs: {json.dumps(needs)}
Stakeholders: {json.dumps(stakeholder)}
Return JSON:
{{"core_problem":"string","problem_tree":{{"immediate_causes":[],"underlying_causes":[],"structural_causes":[],"effects":[]}}}}""")

    individual = call(f"""Perform an INDIVIDUAL & TASK NEEDS ANALYSIS.
Project Data: {fd}
Problem: {json.dumps(problem)}
Return JSON:
{{"task_analysis":[{{"task":"string","complexity":"Low/Medium/High","dependencies":[],"required_competencies":[]}}],"skills_gap_analysis":[{{"role":"string","current_skill_level":"string","required_skill_level":"string","gap":"Low/Medium/High"}}],"training_recommendations":[]}}""")

    objectives = call(f"""Convert the PROBLEM TREE into an OBJECTIVE TREE (negate each negative into a positive objective).
Problem Tree: {json.dumps(problem)}
Return JSON:
{{"end_objective":"string","objective_tree":{{"means_objectives_level_1":[],"means_objectives_level_2":[],"means_objectives_level_3":[]}}}}""")

    strategy = call(f"""Perform a STRATEGY ANALYSIS. Identify 2-3 intervention strategies.
Project Data: {fd}
Objectives: {json.dumps(objectives)}
GDSI: {json.dumps(gdsi)}
Org: {json.dumps(org)}
Return JSON:
{{"possible_strategies":[{{"name":"string","description":"string","feasibility_score":7,"cost_estimate":"Low/Medium/High","sustainability_score":8}}],"recommended_strategy":{{"name":"string","justification":"string","risks":[],"mitigations":[]}}}}""")

    all_out = dict(needs=needs, stakeholder=stakeholder, gdsi=gdsi,
                   organizational=org, problem=problem, individual_task=individual,
                   objectives=objectives, strategy=strategy)

    logframe = call(f"""Generate a LOGICAL FRAMEWORK MATRIX (4 rows: Goal, Outcome, Outputs, Activities; 4 cols: Narrative, Indicators, Means of Verification, Assumptions).
All analyses: {json.dumps(all_out, indent=2)}
Return JSON:
{{"logical_framework":{{"goal":{{"narrative":"string","indicators":[],"means_of_verification":[],"assumptions":[]}},"outcome":{{"narrative":"string","indicators":[],"means_of_verification":[],"assumptions":[]}},"outputs":[{{"narrative":"string","indicators":[],"means_of_verification":[],"assumptions":[]}}],"activities":[{{"narrative":"string","indicators":[],"means_of_verification":[],"assumptions":[]}}]}}}}""")

    return {"status": "success", "analyses": all_out,
            "logical_framework": logframe.get("logical_framework", logframe)}


# ─────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────
SLIDES = ["welcome", "form", "analysis", "results"]

def step_dots(current):
    out = []
    for i in range(len(SLIDES)):
        cls = "step-dot"
        if i < current: cls += " done"
        elif i == current: cls += " active"
        out.append(html.Div(className=cls))
    return out

def mk_list(items):
    if not items: return html.Span("—", style={"color":"var(--muted)"})
    return html.Ul([html.Li(i) for i in items], style={"marginTop":"4px"})

def field(label, hint, fid, placeholder, textarea=False):
    if textarea:
        el = dcc.Textarea(id=fid, placeholder=placeholder, style={"width":"100%","minHeight":"88px"})
    else:
        el = dcc.Input(id=fid, type="text", placeholder=placeholder, style={"width":"100%"})
    return html.Div([
        html.Div(label, className="field-label"),
        html.Div(hint,  className="field-hint"),
        el,
    ])


# ─────────────────────────────────────────────────────────────────
# SLIDES
# ─────────────────────────────────────────────────────────────────
def slide_welcome():
    return html.Div(className="slide", children=[
        html.Div("🌍", style={"fontSize":"4rem","marginBottom":"20px","animation":"pulse 2s ease-in-out infinite"}),
        html.H1(["MEAL ", html.Span("Co-Pilot", className="gradient-text")], className="slide-title"),
        html.P("AI-powered project design and logical framework generation for development organizations.",
               className="slide-subtitle"),
        html.Div(className="stat-strip", children=[
            html.Div([html.Div("9", className="stat-num", style={"color":"var(--accent)"}),
                      html.Div("Analyses", className="stat-lbl")], style={"textAlign":"center"}),
            html.Div(className="stat-divider"),
            html.Div([html.Div("8", className="stat-num", style={"color":"var(--accent2)"}),
                      html.Div("Input Fields", className="stat-lbl")], style={"textAlign":"center"}),
            html.Div(className="stat-divider"),
            html.Div([html.Div("1", className="stat-num", style={"color":"var(--gold)"}),
                      html.Div("LogFrame", className="stat-lbl")], style={"textAlign":"center"}),
        ]),
        html.Button("Begin Analysis →", id="btn-start", className="btn-primary", n_clicks=0),
    ])


def slide_form():
    return html.Div(className="slide", style={"justifyContent":"flex-start","paddingTop":"40px"}, children=[
        html.H2(["Project ", html.Span("Details", className="gradient-text")],
                className="slide-title", style={"fontSize":"2.4rem"}),
        html.P("Fill in the 8 fields to generate your MEAL analysis.",
               className="slide-subtitle", style={"marginBottom":"32px"}),
        html.Div(className="form-card", children=[
            html.Div(className="form-row", children=[
                field("Project Name","Official project name","f-pname","e.g. Rural Footbridge Construction"),
                field("Location","Region, district, village","f-loc","e.g. Western Kenya, Migori County"),
            ]),
            html.Div(className="form-row", children=[
                field("Implementing Organization","Org name, budget, staff count","f-org","e.g. Community CBO, $50k, 3 staff"),
                field("Target Beneficiaries","Number + demographics","f-ben","e.g. 1,200 people, 55% women, 8% PWD"),
            ]),
            html.Div(className="form-row single", children=[
                field("Problem Statement","One sentence — the core problem","f-prob",
                      "e.g. Village cut off during rains; children cannot reach school",textarea=True),
            ]),
            html.Div(className="form-row", children=[
                field("Key Stakeholders","Government, community, NGOs","f-stake",
                      "e.g. County govt, KPLC, local elders, transport union",textarea=True),
                field("Current Situation","Describe the problem with data","f-cur",
                      "e.g. Manual river crossing, 3 deaths last year, 40% absenteeism",textarea=True),
            ]),
            html.Div(className="form-row single", children=[
                field("Desired Situation","How will things look after the project?","f-des",
                      "e.g. Safe all-weather crossing, <5% absenteeism, zero crossing deaths",textarea=True),
            ]),
            html.Div(id="form-error"),
            html.Div([
                html.Button("← Back", id="btn-form-back", className="btn-secondary", n_clicks=0),
                html.Button("Run MEAL Analysis →", id="btn-analyze", className="btn-primary",
                            n_clicks=0, style={"marginLeft":"auto"}),
            ], style={"display":"flex","alignItems":"center","marginTop":"8px","gap":"16px"}),
        ]),
    ])


def slide_analysis():
    return html.Div(className="slide", children=[
        html.Div(className="progress-wrap", children=[
            html.Div("🔬", className="progress-icon"),
            html.H2(["Running ", html.Span("Analysis", className="gradient-text")],
                    className="slide-title", style={"fontSize":"2.2rem"}),
            html.P("GPT is running 9 sequential MEAL analyses. This typically takes 30–90 seconds.",
                   className="slide-subtitle"),
            html.Div(className="progress-bar-bg", children=[
                html.Div(id="pb", className="progress-bar-fill", style={"width":"5%"}),
            ]),
            html.Div(id="pstatus", className="progress-step", children="⚙️ Initializing..."),
        ]),
    ])


# ── Result renderers ──
def r_needs(a):
    n = a.get("needs", {})
    return [
        html.Div("Needs Statement", className="result-label"),
        html.Div(n.get("needs_statement","—"), className="result-value highlight-box"),
        html.Div("Maslow's Hierarchy", className="result-label"),
        html.Div(className="result-value", children=[
            html.B("Physical: "), html.Span(", ".join(n.get("maslow_analysis",{}).get("physical_needs",[]) or ["—"])), html.Br(),
            html.B("Safety: "),   html.Span(", ".join(n.get("maslow_analysis",{}).get("safety_needs",[]) or ["—"])), html.Br(),
            html.B("Belonging: "),html.Span(", ".join(n.get("maslow_analysis",{}).get("social_belonging",[]) or ["—"])), html.Br(),
            html.B("Esteem: "),   html.Span(", ".join(n.get("maslow_analysis",{}).get("esteem_needs",[]) or ["—"])),
        ]),
        html.Div("Prioritized Needs", className="result-label"),
        html.Div([
            html.Div([
                html.Span(f"#{x.get('rank',i+1)} ", style={"color":"var(--gold)","fontWeight":"700"}),
                html.Span(x.get("need",""), style={"fontWeight":"600"}), html.Br(),
                html.Span(x.get("justification",""), style={"color":"var(--muted)","fontSize":"0.82rem"}),
            ], style={"padding":"12px","background":"rgba(79,156,247,0.06)","borderRadius":"8px",
                      "border":"1px solid var(--border)","marginBottom":"8px"})
            for i,x in enumerate(n.get("prioritized_needs",[]))
        ]),
    ]

def r_stakeholder(a):
    sh = a.get("stakeholder", {})
    cm = {"Manage Closely":"#4f9cf7","Keep Satisfied":"#f0c040","Keep Informed":"#7ce0a0","Monitor":"#8ba4c8"}
    return [
        html.Table(className="sh-table", children=[
            html.Thead(html.Tr([html.Th(h) for h in ["Stakeholder","Type","Power","Interest","Grid Position","Engagement"]])),
            html.Tbody([html.Tr([
                html.Td(r.get("name",""), style={"fontWeight":"600"}),
                html.Td(r.get("type","")),
                html.Td(html.Span(r.get("power",""), className=f"badge {'high' if r.get('power')=='High' else 'low'}")),
                html.Td(html.Span(r.get("interest",""), className=f"badge {'high' if r.get('interest')=='High' else 'low'}")),
                html.Td(r.get("grid_position",""), style={"color":cm.get(r.get("grid_position",""),"var(--white)")}),
                html.Td(r.get("engagement_strategy",""), style={"fontSize":"0.8rem","color":"var(--muted)"}),
            ]) for r in sh.get("stakeholder_matrix",[])]),
        ]),
        html.Div("Vulnerable Groups", className="result-label"),
        html.Div(className="badge-grid", children=[html.Span(g, className="badge") for g in sh.get("vulnerable_groups",[])]),
    ]

def r_gdsi(a):
    g = a.get("gdsi", {})
    if "error" in g: return [html.Div(f"Error: {g['error']}", className="error-box")]
    return [
        html.Div("Excluded Groups", className="result-label"),
        html.Div(className="badge-grid", children=[html.Span(x, className="badge") for x in g.get("excluded_groups",[])]),
        html.Div("Barriers Analysis", className="result-label"),
        html.Div([
            html.Div([html.B(b.get("group","")), html.Span(f" · {b.get('barrier_type','')}", style={"color":"var(--accent)","fontSize":"0.8rem"}),
                      html.Br(), html.Span(b.get("description",""), style={"color":"var(--muted)","fontSize":"0.82rem"})],
                     style={"marginBottom":"8px"})
            for b in g.get("barriers_analysis",[])
        ], className="result-value"),
        html.Div("Do No Harm Assessment", className="result-label"),
        html.Div(g.get("do_no_harm_assessment","—"), className="result-value highlight-box"),
        html.Div("Inclusion Actions", className="result-label"),
        html.Div([html.B("Targeted: "), mk_list(g.get("inclusion_actions",{}).get("targeted_actions",[])),
                  html.B("Mainstreamed: ", style={"display":"block","marginTop":"8px"}),
                  mk_list(g.get("inclusion_actions",{}).get("mainstreamed_actions",[]))], className="result-value"),
    ]

def r_problem(a):
    p = a.get("problem", {})
    pt = p.get("problem_tree", {})
    return [
        html.Div("Core Problem", className="result-label"),
        html.Div(p.get("core_problem","—"), className="cyan-box", style={"fontFamily":"Playfair Display,serif","fontSize":"1.1rem","color":"var(--accent2)"}),
        html.Div("Immediate Causes", className="result-label"), html.Div(mk_list(pt.get("immediate_causes",[])), className="result-value"),
        html.Div("Underlying Causes", className="result-label"), html.Div(mk_list(pt.get("underlying_causes",[])), className="result-value"),
        html.Div("Structural Causes", className="result-label"), html.Div(mk_list(pt.get("structural_causes",[])), className="result-value"),
        html.Div("Effects", className="result-label"), html.Div(mk_list(pt.get("effects",[])), className="result-value"),
    ]

def r_objectives(a):
    o = a.get("objectives", {})
    ot = o.get("objective_tree", {})
    return [
        html.Div("End Objective", className="result-label"),
        html.Div(o.get("end_objective","—"), className="gold-box", style={"fontFamily":"Playfair Display,serif","fontSize":"1.1rem","color":"var(--gold)"}),
        html.Div("Level 1 Means Objectives", className="result-label"), html.Div(mk_list(ot.get("means_objectives_level_1",[])), className="result-value"),
        html.Div("Level 2 Means Objectives", className="result-label"), html.Div(mk_list(ot.get("means_objectives_level_2",[])), className="result-value"),
        html.Div("Level 3 Means Objectives", className="result-label"), html.Div(mk_list(ot.get("means_objectives_level_3",[])), className="result-value"),
    ]

def r_strategy(a):
    s = a.get("strategy", {})
    rec = s.get("recommended_strategy", {})
    return [
        html.Div("Strategy Options", className="result-label"),
        html.Div(className="strategy-grid", children=[
            html.Div([
                html.Div(st.get("name",""), style={"fontWeight":"700","marginBottom":"6px"}),
                html.Div(st.get("description",""), style={"fontSize":"0.82rem","color":"var(--muted)","marginBottom":"10px"}),
                html.Div([html.Span("Feasibility ", style={"color":"var(--muted)"}), html.Span(f"{st.get('feasibility_score','?')}/10", style={"color":"var(--accent)"})]),
                html.Div(className="score-bar", children=[html.Div(className="score-fill", style={"width":f"{int(st.get('feasibility_score',0))*10}%"})]),
                html.Div([html.Span("Sustainability ", style={"color":"var(--muted)"}), html.Span(f"{st.get('sustainability_score','?')}/10", style={"color":"var(--accent2)"})], style={"marginTop":"8px"}),
                html.Div(className="score-bar", children=[html.Div(className="score-fill", style={"width":f"{int(st.get('sustainability_score',0))*10}%","background":"linear-gradient(90deg,var(--accent2),var(--gold))"})]),
                html.Span(st.get("cost_estimate",""), className="badge", style={"marginTop":"10px","display":"inline-block"}),
            ], className=f"strategy-card{'  rec' if st.get('name')==rec.get('name') else ''}")
            for st in s.get("possible_strategies",[])
        ]),
        html.Div("Recommended Strategy", className="result-label"),
        html.Div([
            html.Div(rec.get("name","—"), style={"fontWeight":"700","fontSize":"1.1rem","marginBottom":"8px"}),
            html.Div(rec.get("justification",""), style={"color":"var(--muted)","marginBottom":"16px","fontSize":"0.9rem"}),
            html.B("Risks"), mk_list(rec.get("risks",[])),
            html.B("Mitigations", style={"display":"block","marginTop":"8px"}), mk_list(rec.get("mitigations",[])),
        ], className="gold-box"),
    ]

def r_org(a):
    o = a.get("organizational", {})
    sa = o.get("strategic_alignment", {})
    return [
        html.Div("Strategic Alignment", className="result-label"),
        html.Div([
            html.Div(f"Mission: {sa.get('org_mission','—')}", className="result-value"),
            html.Div(f"Contribution: {sa.get('project_contribution','—')}", className="result-value", style={"marginTop":"6px"}),
            html.Span(f"Alignment: {sa.get('alignment_score','—')}", className="badge high", style={"marginTop":"10px","display":"inline-block"}),
        ], className="highlight-box"),
        html.Div("Performance Gaps", className="result-label"),
        html.Div([
            html.Div([html.B(g.get("gap","")), html.Br(),
                      html.Span(f"Root cause: {g.get('root_cause','')}", style={"color":"var(--muted)","fontSize":"0.82rem"}), html.Br(),
                      html.Span(f"Impact: {g.get('impact','')}", style={"color":"#ff9090","fontSize":"0.82rem"})],
                     style={"marginBottom":"10px"})
            for g in o.get("performance_gaps",[])
        ], className="result-value"),
        html.Div(f"Change Readiness: {o.get('change_readiness','—')}", className="result-label"),
    ]

def r_individual(a):
    it = a.get("individual_task", {})
    return [
        html.Div("Task Analysis", className="result-label"),
        html.Div([
            html.Div([
                html.Span(t.get("task",""), style={"fontWeight":"700"}),
                html.Span(f" · {t.get('complexity','')} complexity", style={"color":"var(--muted)","fontSize":"0.8rem"}), html.Br(),
                html.Span("Competencies: "+", ".join(t.get("required_competencies",[])), style={"color":"var(--muted)","fontSize":"0.82rem","marginTop":"4px"}),
            ], style={"padding":"12px","background":"rgba(79,156,247,0.06)","borderRadius":"8px","border":"1px solid var(--border)","marginBottom":"8px"})
            for t in it.get("task_analysis",[])
        ]),
        html.Div("Skills Gap Analysis", className="result-label"),
        html.Table(className="sh-table", children=[
            html.Thead(html.Tr([html.Th(h) for h in ["Role","Current","Required","Gap"]])),
            html.Tbody([html.Tr([
                html.Td(g.get("role",""), style={"fontWeight":"600"}),
                html.Td(g.get("current_skill_level","")),
                html.Td(g.get("required_skill_level","")),
                html.Td(html.Span(g.get("gap",""), className=f"badge {'high' if g.get('gap')=='High' else 'low' if g.get('gap')=='Low' else ''}")),
            ]) for g in it.get("skills_gap_analysis",[])]),
        ]),
        html.Div("Training Recommendations", className="result-label"),
        html.Div(mk_list(it.get("training_recommendations",[])), className="result-value"),
    ]


def build_logframe_excel(lf_data, project_name="Project"):
    """Return base64-encoded xlsx of the logframe."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Logical Framework"

    NAVY="0A1628"; ACCENT="1C5FA8"; WHITE="F0F4FF"; LIGHT="EBF2FF"
    LEVEL_COLORS={"Goal":("1B3A6B","FFFFFF"),"Outcome":("1A5276","FFFFFF"),
                  "Output":("1F618D","FFFFFF"),"Activity":("2874A6","FFFFFF")}

    def fill(c): return PatternFill("solid",fgColor=c)
    def bdr():
        s=Side(style="thin",color="C8D8F0")
        return Border(left=s,right=s,top=s,bottom=s)

    ws.merge_cells("A1:E1")
    t=ws["A1"]
    t.value=f"Logical Framework Matrix — {project_name}"
    t.font=Font(name="Calibri",bold=True,size=14,color=WHITE)
    t.fill=fill(NAVY); t.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.row_dimensions[1].height=38

    for col,hdr in enumerate(["Level","Narrative","Indicators","Means of Verification","Assumptions"],1):
        c=ws.cell(row=2,column=col,value=hdr)
        c.font=Font(name="Calibri",bold=True,size=10,color=WHITE)
        c.fill=fill(ACCENT); c.border=bdr()
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.row_dimensions[2].height=22
    for i,w in enumerate([16,44,36,36,36],1):
        ws.column_dimensions[get_column_letter(i)].width=w

    lf=lf_data or {}
    rows=[]
    for lvl,entry in [("Goal",lf.get("goal",{})),("Outcome",lf.get("outcome",{}))]:
        rows.append((lvl,entry))
    for i,o in enumerate(lf.get("outputs",[]),1): rows.append((f"Output {i}",o))
    for i,a in enumerate(lf.get("activities",[]),1): rows.append((f"Activity {i}",a))

    for ri,(lvl,entry) in enumerate(rows,3):
        lk=lvl.split(" ")[0]
        bg,fg=LEVEL_COLORS.get(lk,(ACCENT,WHITE))
        def cv(items): return "\n".join(f"• {x}" for x in items) if items else "—"
        vals=[lvl, entry.get("narrative","—"),
              cv(entry.get("indicators",[])),
              cv(entry.get("means_of_verification",[])),
              cv(entry.get("assumptions",[]))]
        for ci,val in enumerate(vals,1):
            c=ws.cell(row=ri,column=ci,value=val)
            c.alignment=Alignment(vertical="top",wrap_text=True); c.border=bdr()
            if ci==1:
                c.font=Font(name="Calibri",bold=True,size=10,color=fg)
                c.fill=fill(bg)
                c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            else:
                c.font=Font(name="Calibri",size=10)
                c.fill=fill(LIGHT if ri%2==0 else "FFFFFF")
        n_items=max((len(entry.get(k,[])) for k in ["indicators","means_of_verification","assumptions"]),default=1)
        ws.row_dimensions[ri].height=max(60,n_items*18)

    ws.freeze_panes="B3"
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return base64.b64encode(buf.read()).decode()


def r_logframe(lf_data):
    lf = lf_data if lf_data else {}
    if not lf:
        return [html.Div("No logframe data.", className="error-box")]

    entries=[]
    for lvl,entry in [("Goal",lf.get("goal",{})),("Outcome",lf.get("outcome",{}))]:
        entries.append((lvl,entry))
    for i,o in enumerate(lf.get("outputs",[]),1): entries.append((f"Output {i}",o))
    for i,a in enumerate(lf.get("activities",[]),1): entries.append((f"Activity {i}",a))

    colors={"Goal":"var(--accent)","Outcome":"var(--accent2)","Output":"var(--gold)","Activity":"#7ce0a0"}
    preview=[]
    for lvl,entry in entries:
        lk=lvl.split(" ")[0]
        col=colors.get(lk,"var(--white)")
        preview.append(html.Div([
            html.Div(lvl,style={"fontWeight":"700","fontSize":"0.75rem","letterSpacing":"0.1em",
                                "textTransform":"uppercase","color":col,"marginBottom":"4px"}),
            html.Div(entry.get("narrative","—"),style={"fontSize":"0.9rem","color":"var(--white)","lineHeight":"1.5"}),
            html.Div("Indicators: "+(", ".join(entry.get("indicators",[])) or "—"),
                     style={"fontSize":"0.78rem","color":"var(--muted)","marginTop":"4px"}),
        ],style={"padding":"14px 18px","borderLeft":f"3px solid {col}",
                 "background":"rgba(15,32,64,0.5)","borderRadius":"0 10px 10px 0","marginBottom":"10px"}))

    return [
        html.Div([
            html.Div("📊",style={"fontSize":"2.5rem","marginBottom":"8px"}),
            html.Div("Your Logical Framework Matrix is ready.",style={
                "fontFamily":"Playfair Display,serif","fontSize":"1.2rem","fontWeight":"700","marginBottom":"6px"}),
            html.Div("Full 4×4 matrix (Goal → Outcome → Outputs → Activities) as a formatted Excel file.",
                     style={"color":"var(--muted)","fontSize":"0.88rem","marginBottom":"20px","maxWidth":"480px"}),
            html.Button("⬇ Download LogFrame Excel",id="btn-dl-excel",className="btn-primary",
                        n_clicks=0,style={"fontSize":"1rem","padding":"16px 48px"}),
            dcc.Download(id="download-excel"),
        ],style={"textAlign":"center","padding":"32px","background":"rgba(79,156,247,0.06)",
                 "border":"1px solid var(--border)","borderRadius":"16px","marginBottom":"28px"}),
        html.Div("Framework Preview",className="result-label"),
        html.Div(preview),
    ]



TABS = [
    ("needs","Needs"),("stakeholder","Stakeholders"),("gdsi","GDSI"),
    ("problem","Problem Tree"),("objectives","Objectives"),("strategy","Strategy"),
    ("organizational","Org Analysis"),("individual_task","Task & Skills"),("logframe","LogFrame"),
]

RENDER_MAP = {
    "needs":         ("Needs Analysis",                    r_needs),
    "stakeholder":   ("Stakeholder Analysis",              r_stakeholder),
    "gdsi":          ("Gender, Disability & Social Inclusion", r_gdsi),
    "problem":       ("Problem Tree Analysis",             r_problem),
    "objectives":    ("Objectives Tree",                   r_objectives),
    "strategy":      ("Strategy Analysis",                 r_strategy),
    "organizational":("Organizational Needs Analysis",     r_org),
    "individual_task":("Individual & Task Analysis",       r_individual),
}

def slide_results(result, form_data=None):
    if not result: return html.Div("No results.", className="slide")
    form_data = form_data or {}
    analyses = result.get("analyses", {})
    pname = result.get("project_name","Your Project")
    return html.Div(className="slide", style={"justifyContent":"flex-start","paddingTop":"40px"}, children=[
        html.H2(html.Span("Analysis Complete", className="gradient-text"),
                className="slide-title", style={"fontSize":"2.2rem","marginBottom":"6px"}),

        # ── Project Summary Card ──────────────────────────────────
        html.Div(className="results-wrap", style={"marginBottom":"24px"}, children=[
            html.Div([
                # Left: project identity
                html.Div([
                    html.Div("Project Summary", style={"fontSize":"0.68rem","letterSpacing":"0.14em",
                             "textTransform":"uppercase","color":"var(--accent)","fontWeight":"600","marginBottom":"6px"}),
                    html.Div(pname, style={"fontFamily":"Playfair Display,serif","fontSize":"1.5rem",
                             "fontWeight":"900","color":"var(--white)","lineHeight":"1.2","marginBottom":"10px"}),
                    html.Div([
                        html.Span("📍 ", style={"marginRight":"4px"}),
                        html.Span(result.get("analyses",{}).get("organizational",{}).get("strategic_alignment",{}).get("org_mission",
                                  form_data.get("location","")) or form_data.get("location",""),
                                  style={"color":"var(--muted)","fontSize":"0.88rem"}),
                    ], style={"marginBottom":"6px"}),
                ], style={"flex":"1","minWidth":"200px"}),
                # Divider
                html.Div(style={"width":"1px","background":"var(--border)","margin":"0 24px","alignSelf":"stretch"}),
                # Middle: key stats
                html.Div([
                    html.Div([
                        html.Div(result.get("analyses",{}).get("organizational",{}).get("change_readiness","—"),
                                 style={"fontSize":"1.4rem","fontWeight":"700","color":"var(--accent)"}),
                        html.Div("Change Readiness",style={"fontSize":"0.72rem","color":"var(--muted)","textTransform":"uppercase","letterSpacing":"0.08em"}),
                    ], style={"textAlign":"center","flex":"1"}),
                    html.Div([
                        html.Div(result.get("analyses",{}).get("organizational",{}).get("strategic_alignment",{}).get("alignment_score","—"),
                                 style={"fontSize":"1.4rem","fontWeight":"700","color":"var(--accent2)"}),
                        html.Div("Alignment Score",style={"fontSize":"0.72rem","color":"var(--muted)","textTransform":"uppercase","letterSpacing":"0.08em"}),
                    ], style={"textAlign":"center","flex":"1"}),
                    html.Div([
                        html.Div(str(len(result.get("analyses",{}).get("stakeholder",{}).get("stakeholder_matrix",[]))),
                                 style={"fontSize":"1.4rem","fontWeight":"700","color":"var(--gold)"}),
                        html.Div("Stakeholders",style={"fontSize":"0.72rem","color":"var(--muted)","textTransform":"uppercase","letterSpacing":"0.08em"}),
                    ], style={"textAlign":"center","flex":"1"}),
                ], style={"display":"flex","gap":"16px","flex":"1"}),
                # Divider
                html.Div(style={"width":"1px","background":"var(--border)","margin":"0 24px","alignSelf":"stretch"}),
                # Right: problem & objective
                html.Div([
                    html.Div("Core Problem", style={"fontSize":"0.68rem","letterSpacing":"0.1em",
                             "textTransform":"uppercase","color":"#ff9090","fontWeight":"600","marginBottom":"4px"}),
                    html.Div(result.get("analyses",{}).get("problem",{}).get("core_problem","—"),
                             style={"fontSize":"0.82rem","color":"var(--white)","lineHeight":"1.4","marginBottom":"10px"}),
                    html.Div("End Objective", style={"fontSize":"0.68rem","letterSpacing":"0.1em",
                             "textTransform":"uppercase","color":"var(--gold)","fontWeight":"600","marginBottom":"4px"}),
                    html.Div(result.get("analyses",{}).get("objectives",{}).get("end_objective","—"),
                             style={"fontSize":"0.82rem","color":"var(--white)","lineHeight":"1.4"}),
                ], style={"flex":"2","minWidth":"260px"}),
            ], style={
                "display":"flex","flexWrap":"wrap","gap":"0",
                "padding":"24px 28px",
                "background":"var(--card-bg)",
                "border":"1px solid var(--border)",
                "borderRadius":"16px",
                "backdropFilter":"blur(16px)",
            }),
        ]),
        html.Div(className="results-wrap", children=[
            html.Div(className="tab-nav", id="tab-nav", children=[
                html.Button(lbl, id=f"tab-{k}", className=f"tab-btn{'  active' if k=='logframe' else ''}", n_clicks=0)
                for k,lbl in TABS
            ]),
            html.Div(id="tab-content", className="result-card", children=[
                html.Div("Logical Framework Matrix", className="result-section-title"),
                *r_logframe(result.get("logical_framework",{})),
            ]),
        ]),
        html.Div([
            html.Button("← New Analysis", id="btn-restart", className="btn-secondary", n_clicks=0),
            html.Button("⬇ Download JSON", id="btn-download", className="btn-primary", n_clicks=0),
            dcc.Download(id="download-json"),
        ], style={"display":"flex","gap":"16px","marginTop":"32px","alignItems":"center"}),
    ])


# ─────────────────────────────────────────────────────────────────
# LAYOUT
# ─────────────────────────────────────────────────────────────────
app.layout = html.Div(children=[
    html.Div(className="bg-mesh"),
    html.Div(className="grid-lines"),
    dcc.Store(id="slide", data=0),
    dcc.Store(id="result-store", data=None),
    dcc.Store(id="form-store", data={}),
    dcc.Store(id="running", data=False),
    dcc.Store(id="pct", data=5),
    dcc.Interval(id="pinterval", interval=3000, n_intervals=0, disabled=True),
    html.Div(className="app-header", children=[
        html.Div([
            html.Div("MEAL Co-Pilot", className="logo-mark"),
            html.Div("Momentum Incubator · AI-Powered Project Design", className="logo-sub"),
        ]),
        html.Div(id="dots", className="step-indicator", children=step_dots(0)),
    ]),
    html.Div(id="main", className="slide-wrap", children=[slide_welcome()]),
])


app.validation_layout = html.Div([
    app.layout,
    slide_form(),
    slide_analysis(),
    slide_results({
        "project_name": "Validation Project",
        "analyses": {
            "needs": {},
            "stakeholder": {},
            "gdsi": {},
            "problem": {},
            "objectives": {},
            "strategy": {},
            "organizational": {},
            "individual_task": {},
        },
        "logical_framework": {
            "goal": {"narrative": "", "indicators": [], "means_of_verification": [], "assumptions": []},
            "outcome": {"narrative": "", "indicators": [], "means_of_verification": [], "assumptions": []},
            "outputs": [],
            "activities": [],
        },
    }),
])


# ─────────────────────────────────────────────────────────────────
# CALLBACKS
# ─────────────────────────────────────────────────────────────────

@app.callback(Output("slide","data",allow_duplicate=True), Input("btn-start","n_clicks"), prevent_initial_call=True)
def go_form(n): return 1 if n else no_update

@app.callback(Output("slide","data",allow_duplicate=True), Input("btn-form-back","n_clicks"), prevent_initial_call=True)
def go_welcome(n): return 0 if n else no_update

@app.callback(
    Output("slide","data",allow_duplicate=True), Output("result-store","data",allow_duplicate=True),
    Input("btn-restart","n_clicks"), prevent_initial_call=True,
)
def restart(n): return (0, None) if n else (no_update, no_update)

@app.callback(
    Output("slide","data",allow_duplicate=True),
    Output("form-store","data"),
    Output("form-error","children"),
    Output("running","data"),
    Output("pinterval","disabled"),
    Output("pct","data",allow_duplicate=True),
    Input("btn-analyze","n_clicks"),
    State("f-pname","value"), State("f-loc","value"), State("f-org","value"),
    State("f-ben","value"),   State("f-prob","value"), State("f-stake","value"),
    State("f-cur","value"),   State("f-des","value"),
    prevent_initial_call=True,
)
def start_analysis(n, pname, loc, org, ben, prob, stake, cur, des):
    if not n: return [no_update]*6
    vals = {"project_name":pname,"location":loc,"organization":org,"target_beneficiaries":ben,
            "problem_statement":prob,"stakeholders":stake,"current_situation":cur,"desired_situation":des}
    missing = [k.replace("_"," ").title() for k,v in vals.items() if not v or not str(v).strip()]
    if missing:
        return no_update, no_update, html.Div(f"Please fill in: {', '.join(missing)}", className="error-box"), no_update, no_update, no_update
    return 2, vals, None, True, False, 5

@app.callback(
    Output("result-store","data"),
    Output("running","data",allow_duplicate=True),
    Input("running","data"),
    State("form-store","data"),
    prevent_initial_call=True,
)
def run_bg(running, form_data):
    if not running or not form_data: return no_update, no_update
    try:
        result = run_meal_analysis(form_data)
        result["project_name"] = form_data.get("project_name","")
        return result, False
    except Exception as e:
        return {"status":"error","error":str(e)}, False

STEPS = [
    (12, "🔍 Analyzing community needs..."),
    (26, "👥 Mapping stakeholders..."),
    (40, "♀️ GDSI & organizational analysis..."),
    (54, "🌳 Building problem tree..."),
    (66, "🎯 Mapping objectives..."),
    (78, "📊 Evaluating strategies..."),
    (88, "📋 Generating LogFrame..."),
    (95, "✍️ Finalizing report..."),
]

@app.callback(
    Output("pb","style"), Output("pstatus","children"), Output("pct","data"),
    Input("pinterval","n_intervals"), State("pct","data"), prevent_initial_call=True,
)
def tick(n, pct):
    idx = min(n, len(STEPS)-1)
    new_pct, msg = STEPS[idx]
    return {"width":f"{new_pct}%"}, msg, new_pct

@app.callback(
    Output("slide","data",allow_duplicate=True), Output("pinterval","disabled",allow_duplicate=True),
    Input("result-store","data"), State("slide","data"), prevent_initial_call=True,
)
def on_done(result, slide):
    if result and slide == 2: return 3, True
    return no_update, no_update

@app.callback(
    Output("main","children"), Output("dots","children"),
    Input("slide","data"), State("result-store","data"), State("form-store","data"),
)
def render_slide(slide, result, form_data):
    dots = step_dots(slide)
    if slide == 0: return slide_welcome(), dots
    if slide == 1: return slide_form(), dots
    if slide == 2: return slide_analysis(), dots
    return slide_results(result, form_data or {}), dots

@app.callback(
    Output("tab-content","children"),
    Output("tab-nav","children"),
    [Input(f"tab-{k}","n_clicks") for k,_ in TABS],
    State("result-store","data"),
    prevent_initial_call=True,
)
def switch_tab(*args):
    result = args[-1]
    ctx = callback_context
    if not ctx.triggered or not result: return no_update, no_update
    active = ctx.triggered[0]["prop_id"].split(".")[0].replace("tab-","")
    analyses = result.get("analyses",{})
    lf = result.get("logical_framework",{})
    if active == "logframe":
        title, content = "Logical Framework Matrix", r_logframe(lf)
    else:
        title, fn = RENDER_MAP.get(active, ("","—"))
        content = fn(analyses) if callable(fn) else []
    new_tabs = [html.Button(lbl, id=f"tab-{k}", className=f"tab-btn{'  active' if k==active else ''}", n_clicks=0) for k,lbl in TABS]
    return [html.Div(title, className="result-section-title"), *content], new_tabs

@app.callback(
    Output("download-json","data"),
    Input("btn-download","n_clicks"),
    State("result-store","data"),
    prevent_initial_call=True,
)
def dl(n, result):
    if n and result: return dcc.send_string(json.dumps(result, indent=2), "meal_analysis.json")
    return no_update


@app.callback(
    Output("download-excel","data"),
    Input("btn-dl-excel","n_clicks"),
    State("result-store","data"),
    prevent_initial_call=True,
)
def dl_excel(n, result):
    if n and result:
        lf = result.get("logical_framework", {})
        pname = result.get("project_name","Project")
        b64 = build_logframe_excel(lf, pname)
        filename = pname.replace(" ","_")[:40] + "_logframe.xlsx"
        return dict(content=b64, filename=filename, type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", base64=True)
    return no_update

if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=int(os.environ.get("PORT", 8050)))
